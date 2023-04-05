# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

from os import listdir,mkdir,startfile
from os.path import isfile, join,exists
import pandas as pd
from openpyxl import Workbook
from openpyxl import worksheet
import datetime
import time

from openpyxl import load_workbook as lw
from openpyxl.utils import get_column_letter
import openpyxl as xl





path = 'Input'
pdffiles = [f for f in listdir(path) if isfile(join(path, f)) and '.xlsx' in f]
print('\nList of Excel Files:\n')
for file in pdffiles:
    print(file)

    # modified
    wb = lw(path+"\\"+file)
    wb.create_sheet("PR")

    ws = wb["PRList"]
    ws_des = wb["PR"]
    last_empty_row = len(list(ws.rows)) + 1
    print(last_empty_row)
    ref_no = ws['A4'].value

    head_col = 6
    ws['K6'] = 'PR No.'
    for x in range(last_empty_row - head_col-1):
        a = 'K%s' % (x+head_col+1)
        ws[a] = ref_no

    # calculate total number of rows and
    # columns in source excel file
    mr = last_empty_row - head_col-1
    mc = 12

    # copying the cell values from source
    # excel file to destination excel file
    for i in range(head_col, last_empty_row):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws_des.cell(row=i-head_col+1, column=j).value = c.value

    # delete col 1-5
    # ws.delete_rows(1,4)
    file_name = "Input_modified\\" + file
    wb.remove(ws)
    wb.save(filename=file_name)



excl_list = []
path = 'Input_modified'
pdffiles = [f for f in listdir(path) if isfile(join(path, f)) and '.xlsx' in f]
print('\nList of Excel Files:\n')
for file in pdffiles:
    print(file)
    pd_file = pd.read_excel(path + "\\" + file)
    excl_list.append(pd_file)

excl_merged = pd.concat(excl_list)

excl_merged.to_excel('merged.xlsx', index=False)